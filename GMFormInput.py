import openpyxl

class GMFormInput():
	"""
	This class gets the important information for a row in the Input spreadsheet.
	"""
	def __init__(self):
		super(GMFormInput, self).__init__()

	def GMFormInput(self, row, filename):
		loc = filename		
		
		wb = openpyxl.load_workbook(loc)		#open excel sheet
		sheet = wb.active
		#sheet = wb.sheet_by_index(0)		
		# pathname = sheet.cell_value(row,4)     	#Takes each cell from the sheet
		scenario = sheet.cell(row,2).value
		subgoal = sheet.cell(row,3).value
		action = sheet.cell(row,4).value
		url = sheet.cell(row, 5).value   

		# scenario = sheet.cell_value(row,2)
		# subgoal = sheet.cell_value(row,3)
		# action = sheet.cell_value(row,4)
		# url = sheet.cell_value(row, 5)   
		return scenario.lower(), subgoal.lower(), action.lower(), url

	# Fetch "before" action for the current URL
	def before_action(self, row, filename):
		loc = filename		
		
		wb = openpyxl.load_workbook(loc)		#open excel sheet
		sheet = wb.active				
		action = sheet.cell(row,4).value   
		return action.lower()

	# Fetch "after" action for the current URL
	def before_pathname(self, row):
		loc = ('ID_input.xlsx')			
		
		wb = openpyxl.load_workbook(loc)	#open excel sheet
		sheet = wb.active		
		action = sheet.cell(row,4).value   
		return action.lower()

	# Fetch keywords for the "before" action for the current URL
	def before_action_keywords(self, row):
		loc = ('ID_input.xlsx')			
		
		wb = openpyxl.load_workbook(loc)		#open excel sheet
		sheet = wb.active			
		keywords = sheet.cell(row,6).value   
		return keywords
