from GMFormInput import GMFormInput
from parseDOM import parseDOM
from CheckRules import CheckRules
from gensim.summarization import keywords
import re
from openpyxl import load_workbook
import spacy
from collections import Counter
import xlrd

# This file runs each of the rules for each URL in the input/output file

filename = "Example_Input.xlsx"      #Name of the file to take input from
report = open('report.txt', 'w')  #Output file with the reports

#Loading the English model for spaCy
nlp = spacy.load('en_core_web_sm')

# Initialize an object for the classes imported from other files
G = GMFormInput()
P = parseDOM()
C = CheckRules()

res = 0

# List of DOM words to exclude from keywords
DOM_words = ['window', 'document', 'header', 'form', 'link', 'field', 'tab', 'button', 'checkbox', 'icon', 'data', 'information']

for row in range(1, 5): #The number of rows in the input file to run AID on
    keywords_S = []
    keywords_A = []
    all_keywords = []
    usecase, subgoal, action, url = G.GMFormInput(row, filename) # pathname, scenario, subgoal, action, url = G.GMFormInput(row)
    print(row)
    print(url)

    workbook = load_workbook(filename)
    sheet = workbook.active

# Getting Keywords from subgoal and action by extracting nouns
    subgoals = nlp(subgoal)
    actions = nlp(action)

    for token in subgoals:
       if (token.pos_ == 'NOUN' or token.pos_ == 'ADJ') and (str(token) not in DOM_words): #or (token.pos_ == 'NOUN'):
           keywords_S.append(token.text)

    for token in actions:
        if (token.pos_ == 'NOUN' or token.pos_ == 'ADJ') and (str(token) not in DOM_words):
            keywords_A.append(token.text)

    # print(keywords_S,keywords_A)

    P.get_html(url)
    pathname="current_html.html"
    report.write('URL of webpage evaluated: ')
    report.write(url)
    report.write('\n Subgoal: ')
    report.write(subgoal)
    report.write('\n Action: ')
    report.write(action)
    
   # Rule 1 starts here
    result_1_S = C.checkRule1(pathname, keywords_S, 'Alltext.txt')
    result_1_A = C.checkRule1(pathname, keywords_A, 'Alltext.txt')
    if (result_1_S==1 and result_1_A==1):
        sheet.cell(row+1, 6).value = 0 #"ok"
        report.write("\n Rule 1 not violated.")
    else:
        sheet.cell(row+1, 6).value = 1 #"violated"
        report.write("\n Rule 1 violated: Keywords not found on the webpage.")

    print("Rule 1")

    #Rule 2 starts here
    if (row%2==0):
        result_2 = C.checkRule2(pathname, row, filename)
        sheet.cell(row+1, 7).value = result_2
        if result_2==1:
            report.write("\n Rule 2 is violated: Keywords from link-label is not present on the current page.")
        else:
            report.write("\n Rule 2 not violated.")
    else:
        sheet.cell(row+1, 7).value = "N/A"
        report.write("\n Rule 2 not applicable.")

    print("Rule 2")


    # #Rule 3 starts here
    result_3 = C.checkRule3(pathname)
    if result_3==0:
        sheet.cell(row+1, 8).value = 0
        report.write("\n Rule 3 not violated.")
    else:
        sheet.cell(row+1, 8).value = 1
        report.write("\n Rule 3 violated: Link is not labelled.")
    print("Rule 3")
     

    #  #Rule 5 and 6 start here
    results_4 = C.checkRule4(url)
    # print(results_5, results_6, url)
    if (results_4 == -999):
        sheet.cell(row+1, 9).value = 0 #N/A 
        sheet.cell(row+1, 10).value = 0 #N/A
        report.write("\n Rule 4 and 5 not applicable.")
    elif (results_4 == 0):
        sheet.cell(row+1, 9).value = 1
        sheet.cell(row+1, 10).value = 1
        report.write("\n Rule 4 and 5 violated: no labels on issues. ")
    elif results_4>0:
        report.write("\n Rule 4 not violated.")
        results_5 = C.checkRule5(url)
        sheet.cell(row+1, 9).value = 0
        if (results_5 == 0):
            sheet.cell(row+1, 10).value = 1
            report.write("\n Rule 5 violated: no newcomer labels on issues. ")
        elif (results_5 > 0):
            sheet.cell(row+1, 10).value = 0
            report.write("\n Rule 5 not violated. ")
    
    print("Rule 4 + 5")
    workbook.save(filename)
    report.write("\n\n")

report.close()
